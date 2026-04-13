# Module: generator.py

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

def generate_sites_excel(site_df, template_path, output_path):
    """Génère un fichier Excel NOUVEAU contenant UNIQUEMENT les données SITES transformées."""
    
    # Créer un nouveau workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut et créer la feuille 3G
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    ws = wb.create_sheet('3G')
    print(f"✅ Nouvelle feuille 3G créée (vierge)")
    
    print(f"📝 Écriture des SITES dans le nouveau fichier")
    
    # Écrire les en-têtes
    for col_idx, col_name in enumerate(site_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    print(f"   Entêtes sites écrites à la ligne 1")
    
    # Écrire les données sites
    for row_idx, (_, row_data) in enumerate(site_df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    print(f"   ✅ {len(site_df)} sites écrits (lignes 2-{len(site_df)+1})")
    
    # Sauvegarder le fichier
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    print(f"✅ Fichier SITES créé (nouveau): {output_path}")
    
    return output_path


def generate_cells_excel(cell_df, template_path, output_path):
    """Génère un fichier Excel NOUVEAU contenant les données CELLULES (orphelines supprimées)."""
    
    # Créer un nouveau workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut et créer la feuille 3G
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    ws = wb.create_sheet('3G')
    print(f"✅ Nouvelle feuille 3G créée (vierge)")
    
    print(f"📝 Écriture des CELLULES dans le nouveau fichier")
    
    # Définir les couleurs pour les en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Bleu foncé
    header_font = Font(color="FFFFFF", bold=True)
    
    # Écrire les en-têtes
    for col_idx, col_name in enumerate(cell_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    print(f"   Entêtes cellules écrites à la ligne 1")
    
    # Écrire les données cellules
    for row_idx, (_, row_data) in enumerate(cell_df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajuster la largeur des colonnes
    for col_idx in range(1, len(cell_df.columns) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20
    
    print(f"   ✅ {len(cell_df)} cellules écrites (lignes 2-{len(cell_df)+1})")
    
    # Sauvegarder le fichier
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    print(f"✅ Fichier CELLULES créé (nouveau): {output_path}")
    
    return output_path




def generate_final_excel(site_df, cell_df, output_path_sites, output_path_cells):
    """Génère deux fichiers Excel neufs: un pour les SITES et un pour les CELLULES (données transformées uniquement)."""
    
    print(f"🔄 Génération des deux fichiers de sortie...")
    
    # Générer fichier SITES
    print(f"\n📍 Génération fichier SITES...")
    generate_sites_excel(site_df, None, output_path_sites)
    
    # Générer fichier CELLULES
    print(f"\n📍 Génération fichier CELLULES...")
    generate_cells_excel(cell_df, None, output_path_cells)
    
    print(f"\n✅ Deux fichiers créés (neufs) avec succès!")
    
    return output_path_sites, output_path_cells


def generate_gsm_sites_excel(site_df, template_path, output_path):
    """Génère le fichier SITES pour 2G GSM."""
    print(f"📝 Création workbook SITES GSM 2G...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SITES"
    
    # Ajouter les en-têtes avec formatage bleu
    headers = ['GSM_Site', 'nom site', 'code site', 'lat', 'long']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter les données
    for row_idx, row in enumerate(site_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajuster les largeurs de colonnes
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    wb.save(output_path)
    print(f"   ✅ Fichier SITES GSM créé: {output_path}")


def generate_gsm_cells_excel(cell_df, template_path, output_path):
    """Génère le fichier CELLULES pour 2G GSM."""
    print(f"📝 Création workbook CELLULES GSM 2G...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CELLS"
    
    # En-têtes avec formatage bleu
    headers = ['GSM_Cell', 'code site', 'nom site', 'nom cellule', 'cellId', 'BCCH Frequency']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter les données
    for row_idx, row in enumerate(cell_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajuster les largeurs
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    wb.save(output_path)
    print(f"   ✅ Fichier CELLULES GSM créé: {output_path}")


def generate_gsm_excel(site_df, cell_df, output_path_sites, output_path_cells):
    """Orchestre la génération des fichiers GSM 2G."""
    print(f"\n🎯 Génération fichiers 2G GSM...")
    
    # Générer fichier SITES
    print(f"\n📍 Génération fichier SITES...")
    generate_gsm_sites_excel(site_df, None, output_path_sites)
    
    # Générer fichier CELLULES
    print(f"\n📍 Génération fichier CELLULES...")
    generate_gsm_cells_excel(cell_df, None, output_path_cells)
    
    print(f"\n✅ Deux fichiers GSM 2G créés avec succès!")
    
    return output_path_sites, output_path_cells

